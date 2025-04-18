import streamlit as st
import pandas as pd
import numpy as np
import io
import base64
import time
from rapidfuzz import fuzz, process
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import plotly.graph_objects as go

# Set page config
st.set_page_config(
    page_title="Data Matcher App",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Add custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2rem !important;
        font-weight: 600;
        color: #ff4b4b;
        margin-bottom: 1rem;
    }
    .section-header {
        font-size: 1.5rem !important;
        font-weight: 500;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
    }
    .stButton>button {
        background-color: #ff4b4b;
        color: white;
        font-weight: 500;
    }
    .stButton>button:hover {
        background-color: #ff6b6b;
    }
    .connection-line {
        stroke: #2196F3;
        stroke-width: 2;
    }
    .field-card {
        border: 1px solid #dddddd;
        border-radius: 5px;
        padding: 10px;
        margin-bottom: 10px;
        background-color: white;
    }
    .highlight {
        background-color: #e6f7ff;
        border-left: 3px solid #1890ff;
    }
    .stProgress > div > div > div > div {
        background-color: #ff4b4b;
    }
    .status-text {
        margin-top: 0.5rem;
        margin-bottom: 0.5rem;
        color: #4a4a4a;
        font-size: 1rem;
    }
    .metric-card {
        background-color: #f8f9fa;
        border-radius: 5px;
        padding: 1rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    .download-button {
        display: inline-block;
        background-color: #ff4b4b;
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 5px;
        text-decoration: none;
        font-weight: 500;
        margin-top: 1rem;
    }
    .download-button:hover {
        background-color: #ff6b6b;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# Helper functions for file handling
def load_data(file):
    try:
        # Get file extension
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            df = pd.read_csv(file)
        elif file_extension in ['xls', 'xlsx']:
            df = pd.read_excel(file)
        else:
            st.error(f"Unsupported file format: {file_extension}")
            return None
        
        return df
    except Exception as e:
        st.error(f"Error loading file: {e}")
        return None

def get_download_link(df, filename="matched_results.xlsx"):
    """Generate a download link for the DataFrame as an Excel file"""
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Get ExcelWriter object
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write data to Excel
        df.to_excel(writer, index=False, sheet_name='Matched Results')
        
        # Access workbook and active worksheet
        workbook = writer.book
        worksheet = writer.sheets['Matched Results']
        
        # Format header row
        header_fill = PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_border = Border(
            bottom=Side(style='medium', color="DDDDDD")
        )
        
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            column_width = max(len(str(column)), df[column].astype(str).map(len).max())
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(column_width + 2, 30)
        
        # Zebra striping for rows
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_fill
    
    # Get binary data
    data = output.getvalue()
    
    # Encode to base64
    b64 = base64.b64encode(data).decode()
    
    # Generate download link
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" class="download-button">Download Excel File</a>'
    
    return href

# Function to calculate similarity between two strings
def calculate_similarity(str1, str2, method="jaro_winkler"):
    """Calculate similarity between two strings using specified method"""
    if pd.isna(str1) or pd.isna(str2):
        return 0.0
    
    str1 = str(str1)
    str2 = str(str2)
    
    if method == "levenshtein":
        return fuzz.ratio(str1, str2) / 100.0
    elif method == "jaro_winkler":
        return fuzz.token_sort_ratio(str1, str2) / 100.0
    elif method == "partial_ratio":
        return fuzz.partial_ratio(str1, str2) / 100.0
    elif method == "token_set_ratio":
        return fuzz.token_set_ratio(str1, str2) / 100.0
    else:
        return fuzz.ratio(str1, str2) / 100.0

# Function to match records based on field mappings
def match_datasets(df1, df2, field_mappings, threshold=0.8, method="jaro_winkler", progress_bar=None, status_text=None):
    """Match records between two datasets using fuzzy matching"""
    # Initialize results dataframe
    results = pd.DataFrame()
    
    # Calculate total number of comparisons
    total_comparisons = len(df1) * len(df2)
    processed_comparisons = 0
    
    # Initialize progress tracking
    if progress_bar:
        progress_bar.progress(0)
    if status_text:
        status_text.text(f"Processing: 0/{total_comparisons} comparisons (0%)")
    
    # For each record in df1
    for idx1, row1 in df1.iterrows():
        # For each record in df2
        for idx2, row2 in df2.iterrows():
            # Calculate similarity for each field mapping
            similarities = []
            for field1, field2 in field_mappings:
                sim = calculate_similarity(row1[field1], row2[field2], method)
                similarities.append(sim)
            
            # Calculate average similarity across all field mappings
            if similarities:
                avg_similarity = sum(similarities) / len(similarities)
                
                # If similarity is above threshold, add to results
                if avg_similarity >= threshold:
                    # Create a new row with data from both datasets
                    new_row = {}
                    
                    # Add fields from df1
                    for field in df1.columns:
                        new_row[f"{field}_1"] = row1[field]
                    
                    # Add fields from df2
                    for field in df2.columns:
                        new_row[f"{field}_2"] = row2[field]
                    
                    # Add similarity score
                    new_row["similarity_score"] = avg_similarity
                    
                    # Append to results dataframe
                    results = pd.concat([results, pd.DataFrame([new_row])], ignore_index=True)
            
            # Update progress
            processed_comparisons += 1
            if processed_comparisons % max(1, total_comparisons // 100) == 0 or processed_comparisons == total_comparisons:
                progress_percent = processed_comparisons / total_comparisons
                if progress_bar:
                    progress_bar.progress(progress_percent)
                if status_text:
                    status_text.text(f"Processing: {processed_comparisons}/{total_comparisons} comparisons ({int(progress_percent*100)}%)")
    
    # Update progress to complete
    if progress_bar:
        progress_bar.progress(1.0)
    if status_text:
        status_text.text(f"Completed: {total_comparisons}/{total_comparisons} comparisons (100%)")
    
    # Sort results by similarity score (descending)
    if not results.empty:
        results = results.sort_values(by="similarity_score", ascending=False)
    
    return results

# Initialize session state variables
if 'df1' not in st.session_state:
    st.session_state.df1 = None
if 'df2' not in st.session_state:
    st.session_state.df2 = None
if 'field_mappings' not in st.session_state:
    st.session_state.field_mappings = []
if 'output_fields_1' not in st.session_state:
    st.session_state.output_fields_1 = []
if 'output_fields_2' not in st.session_state:
    st.session_state.output_fields_2 = []
if 'results' not in st.session_state:
    st.session_state.results = None
if 'threshold' not in st.session_state:
    st.session_state.threshold = 0.8
if 'method' not in st.session_state:
    st.session_state.method = "jaro_winkler"

# Main app layout
st.markdown("<h1 class='main-header'>Data Matcher App</h1>", unsafe_allow_html=True)
st.write("Upload two datasets, map fields for comparison, and generate fuzzy matching results.")

# Create sidebar for data upload and matching configuration
with st.sidebar:
    st.markdown("<h2 class='section-header'>1. Upload Data</h2>", unsafe_allow_html=True)
    
    # File uploads
    file1 = st.file_uploader("Upload Dataset 1", type=["csv", "xlsx", "xls"])
    file2 = st.file_uploader("Upload Dataset 2", type=["csv", "xlsx", "xls"])
    
    # Load data when files are uploaded
    if file1 is not None and file2 is not None:
        df1 = load_data(file1)
        df2 = load_data(file2)
        
        if df1 is not None and df2 is not None:
            st.session_state.df1 = df1
            st.session_state.df2 = df2
            
            # Initialize output fields with all columns
            if not st.session_state.output_fields_1:
                st.session_state.output_fields_1 = list(df1.columns)
            if not st.session_state.output_fields_2:
                st.session_state.output_fields_2 = list(df2.columns)
    
    st.markdown("<h2 class='section-header'>3. Configure Matching</h2>", unsafe_allow_html=True)
    
    # Similarity threshold slider
    threshold = st.slider(
        "Similarity Threshold",
        min_value=0.0,
        max_value=1.0,
        value=st.session_state.threshold,
        step=0.05,
        format="%.2f",
        help="Minimum similarity score required for a match (0 = no similarity, 1 = exact match)"
    )
    st.session_state.threshold = threshold
    
    # Matching method selection
    method = st.selectbox(
        "Matching Method",
        options=["jaro_winkler", "levenshtein", "partial_ratio", "token_set_ratio"],
        index=["jaro_winkler", "levenshtein", "partial_ratio", "token_set_ratio"].index(st.session_state.method),
        help="Algorithm used to calculate string similarity"
    )
    st.session_state.method = method
    
    # Process button
    process_button = st.button("Process Data", type="primary")

# Main content area
if st.session_state.df1 is not None and st.session_state.df2 is not None:
    # Display data previews
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("<h3>Dataset 1 Preview</h3>", unsafe_allow_html=True)
        st.dataframe(st.session_state.df1.head(5), use_container_width=True)
    
    with col2:
        st.markdown("<h3>Dataset 2 Preview</h3>", unsafe_allow_html=True)
        st.dataframe(st.session_state.df2.head(5), use_container_width=True)
    
    # Field mapping section
    st.markdown("<h2 class='section-header'>2. Map Fields for Comparison</h2>", unsafe_allow_html=True)
    st.write("Connect fields from Dataset 1 to corresponding fields in Dataset 2 for comparison.")
    
    # Create columns for field mapping
    map_col1, map_col2, map_col3 = st.columns([2, 1, 2])
    
    field_mappings = []
    
    with map_col1:
        st.markdown("<h4>Dataset 1 Fields</h4>", unsafe_allow_html=True)
        fields1 = st.session_state.df1.columns.tolist()
        selected_fields1 = []
        
        for field in fields1:
            if st.checkbox(field, key=f"field1_{field}"):
                selected_fields1.append(field)
    
    with map_col3:
        st.markdown("<h4>Dataset 2 Fields</h4>", unsafe_allow_html=True)
        fields2 = st.session_state.df2.columns.tolist()
        selected_fields2 = []
        
        for field in fields2:
            if st.checkbox(field, key=f"field2_{field}"):
                selected_fields2.append(field)
    
    # Display field mappings
    with map_col2:
        st.markdown("<h4>Mappings</h4>", unsafe_allow_html=True)
        
        if selected_fields1 and selected_fields2:
            st.write("Create mapping:")
            field1 = st.selectbox("Dataset 1 Field", options=selected_fields1, key="mapping_field1")
            field2 = st.selectbox("Dataset 2 Field", options=selected_fields2, key="mapping_field2")
            
            if st.button("Add Mapping", key="add_mapping"):
                new_mapping = (field1, field2)
                if new_mapping not in st.session_state.field_mappings:
                    st.session_state.field_mappings.append(new_mapping)
    
    # Display current mappings
    if st.session_state.field_mappings:
        st.markdown("<h4>Current Field Mappings</h4>", unsafe_allow_html=True)
        
        # Create a Plotly figure for visualizing the mappings
        fig = go.Figure()
        
        # Add horizontal lines for connections
        y_pos = 1
        for field1, field2 in st.session_state.field_mappings:
            fig.add_shape(
                type="line",
                x0=0.2, y0=y_pos,
                x1=0.8, y1=y_pos,
                line=dict(color="#2196F3", width=2)
            )
            
            # Add field names
            fig.add_annotation(
                x=0.1, y=y_pos,
                text=field1,
                showarrow=False,
                xanchor="right"
            )
            
            fig.add_annotation(
                x=0.9, y=y_pos,
                text=field2,
                showarrow=False,
                xanchor="left"
            )
            
            # Add delete button
            fig.add_annotation(
                x=0.5, y=y_pos,
                text="❌",
                showarrow=False,
                clicktoshow=False
            )
            
            y_pos += 1
        
        # Configure figure layout
        fig.update_layout(
            height=50 * len(st.session_state.field_mappings) + 50,
            showlegend=False,
            xaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
            yaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
            margin=dict(l=0, r=0, t=10, b=10)
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Button to clear all mappings
        if st.button("Clear All Mappings"):
            st.session_state.field_mappings = []
            st.rerun()
        
        # Create columns for the mappings
        for i, (field1, field2) in enumerate(st.session_state.field_mappings):
            cols = st.columns([4, 1, 4, 1])
            cols[0].write(field1)
            cols[1].write("➡️")
            cols[2].write(field2)
            if cols[3].button("🗑️", key=f"delete_{i}"):
                st.session_state.field_mappings.pop(i)
                st.rerun()
    
    # Output fields selection
    st.markdown("<h2 class='section-header'>3. Select Output Fields</h2>", unsafe_allow_html=True)
    st.write("Choose which fields to include in the final results file.")
    
    output_col1, output_col2 = st.columns(2)
    
    with output_col1:
        st.markdown("<h4>Dataset 1 Output Fields</h4>", unsafe_allow_html=True)
        
        output_fields_1 = []
        for field in st.session_state.df1.columns:
            if st.checkbox(field, key=f"output1_{field}", value=field in st.session_state.output_fields_1):
                output_fields_1.append(field)
        
        st.session_state.output_fields_1 = output_fields_1
    
    with output_col2:
        st.markdown("<h4>Dataset 2 Output Fields</h4>", unsafe_allow_html=True)
        
        output_fields_2 = []
        for field in st.session_state.df2.columns:
            if st.checkbox(field, key=f"output2_{field}", value=field in st.session_state.output_fields_2):
                output_fields_2.append(field)
        
        st.session_state.output_fields_2 = output_fields_2
    
    # Process the data when button is clicked
    if process_button:
        if not st.session_state.field_mappings:
            st.error("Please create at least one field mapping before processing.")
        else:
            # Create progress indicators
            st.markdown("<h3>Processing Data</h3>", unsafe_allow_html=True)
            progress_bar = st.progress(0)
            status_text = st.empty()
            stats_container = st.container()
            
            # Display processing stats
            with stats_container:
                stats_cols = st.columns(4)
                record_count1 = stats_cols[0].metric("Dataset 1 Records", len(st.session_state.df1))
                record_count2 = stats_cols[1].metric("Dataset 2 Records", len(st.session_state.df2))
                total_comparisons = stats_cols[2].metric("Total Comparisons", len(st.session_state.df1) * len(st.session_state.df2))
                est_time = stats_cols[3].metric("Est. Processing Time", f"{len(st.session_state.df1) * len(st.session_state.df2) // 5000 + 1}s")
            
            # Start a timer
            start_time = time.time()
            
            # Perform matching with progress tracking
            results = match_datasets(
                st.session_state.df1,
                st.session_state.df2,
                st.session_state.field_mappings,
                threshold=st.session_state.threshold,
                method=st.session_state.method,
                progress_bar=progress_bar,
                status_text=status_text
            )
            
            # Calculate processing time
            processing_time = time.time() - start_time
            
            # Update the status with actual processing time
            status_text.text(f"Completed in {processing_time:.2f} seconds. Found {len(results)} matches.")
            
            # Filter results to only include selected output fields
            if not results.empty:
                output_columns = []
                
                # Add selected fields from df1
                for field in st.session_state.output_fields_1:
                    output_columns.append(f"{field}_1")
                
                # Add selected fields from df2
                for field in st.session_state.output_fields_2:
                    output_columns.append(f"{field}_2")
                
                # Add similarity score
                output_columns.append("similarity_score")
                
                # Filter columns
                results = results[output_columns]
            
            st.session_state.results = results
    
    # Display results
    if st.session_state.results is not None:
        st.markdown("<h2 class='section-header'>4. Matching Results</h2>", unsafe_allow_html=True)
        
        if st.session_state.results.empty:
            st.warning("No matches found with the current threshold. Try lowering the threshold value.")
        else:
            st.write(f"Found {len(st.session_state.results)} matches.")
            st.dataframe(st.session_state.results, use_container_width=True)
            
            # Add download button
            st.markdown(get_download_link(st.session_state.results), unsafe_allow_html=True)

else:
    # Show welcome message if no data is uploaded yet
    st.info("👆 Please upload two datasets using the sidebar to get started.")
    
    # Show example usage
    st.markdown("<h2 class='section-header'>How to Use This App</h2>", unsafe_allow_html=True)
    
    st.markdown("""
    1. **Upload Data**: Start by uploading two datasets (CSV or Excel files) you want to compare.
    2. **Map Fields**: Create connections between related fields in both datasets that should be compared.
    3. **Configure Matching**: Set the similarity threshold and matching method in the sidebar.
    4. **Select Output Fields**: Choose which fields from each dataset should appear in the results.
    5. **Process Data**: Click the "Process Data" button to run the matching algorithm.
    6. **Review Results**: Examine the matching results and download them as an Excel file.
    """)
    
    st.markdown("<h2 class='section-header'>Example Use Cases</h2>", unsafe_allow_html=True)
    
    st.markdown("""
    - **Customer Database Deduplication**: Find duplicate customer records across different systems.
    - **Address Verification**: Match address records from different sources.
    - **Product Catalog Matching**: Compare product listings from different suppliers.
    - **Employee Record Reconciliation**: Match employee records across HR systems.
    - **Contact List Consolidation**: Combine contact lists while identifying duplicates.
    """)

# Add footer
st.markdown("""
---
<p style="text-align: center; color: #888888;">
    Data Matcher App © 2025
</p>
""", unsafe_allow_html=True)